import csv
import os
from datetime import datetime
import openpyxl
from django.conf import settings

from .models import Export
from apps.companies.models import Company

EXPORT_HEADERS = [
    "ID", "Company Name", "Founders", "Sector", "Current Stage",
    "Team Size", "Member Since", "Key Highlights", "About",
    "Website", "LinkedIn", "Email", "Phone", "Location",
    "Engagement Level", "Smart Card Number", "Startup Type",
    "Ecosystem Category", "Team Members", "Source URL", "Logo URL", "Scraped At",
]


def _company_row(c):
    return [
        c.id, c.company_name, c.founders, c.sector, c.current_stage,
        c.team_size, c.member_since, c.key_highlights, c.about,
        c.website, c.linkedin, c.email, c.phone, c.location,
        c.engagement_level, c.smart_card_number, c.startup_type, c.ecosystem_category, c.team_members,
        c.profile_url, c.logo_url,
        c.scraped_at.strftime("%Y-%m-%d %H:%M:%S") if c.scraped_at else "",
    ]


class ExportService:
    @staticmethod
    def generate_excel(job_id=None, user=None, queryset=None) -> str:
        export_dir = getattr(settings, 'EXPORTS_PATH', str(settings.BASE_DIR / 'exports'))
        os.makedirs(export_dir, exist_ok=True)

        companies = queryset if queryset is not None else Company.objects.all().order_by('-scraped_at')
        if not companies.exists():
            raise ValueError("No companies available for export")

        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"StartupTN_Companies_{date_str}.xlsx"
        file_path = os.path.join(export_dir, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Companies"

        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill("solid", fgColor="1976D2")
        ws.append(EXPORT_HEADERS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for c in companies:
            ws.append(_company_row(c))

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        wb.save(file_path)

        Export.objects.create(
            filename=filename,
            file_type='excel',
            file_path=file_path,
            file_size=os.path.getsize(file_path),
            total_records=companies.count(),
            job_id=job_id,
            created_by=user if user and user.is_authenticated else None,
        )

        return file_path

    @staticmethod
    def generate_csv(job_id=None, user=None, queryset=None) -> str:
        export_dir = getattr(settings, 'EXPORTS_PATH', str(settings.BASE_DIR / 'exports'))
        os.makedirs(export_dir, exist_ok=True)

        companies = queryset if queryset is not None else Company.objects.all().order_by('-scraped_at')
        if not companies.exists():
            raise ValueError("No companies available for export")

        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"StartupTN_Companies_{date_str}.csv"
        file_path = os.path.join(export_dir, filename)

        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(EXPORT_HEADERS)
            for c in companies:
                writer.writerow(_company_row(c))

        Export.objects.create(
            filename=filename,
            file_type='csv',
            file_path=file_path,
            file_size=os.path.getsize(file_path),
            total_records=companies.count(),
            job_id=job_id,
            created_by=user if user and user.is_authenticated else None,
        )

        return file_path
